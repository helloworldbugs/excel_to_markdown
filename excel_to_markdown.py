import pandas as pd
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import filedialog

pwd = os.getcwd()

root = tk.Tk()
root.withdraw()

# 获取选择好的文件
filepath = filedialog.askopenfilename(title="请选择Excel文件", filetypes=[("Excel files", "*.xlsx;*.xls")])



def excel_to_markdown_html(input_file, output_file):

    # 使用openpyxl加载工作簿
    wb = load_workbook(input_file)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # 读取Excel文件
            df = pd.read_excel(input_file, sheet_name=sheet_name, header=None, engine='openpyxl')
            
            f.write('<table>\n')

            # 获取合并单元格信息
            merged_cells = sheet.merged_cells.ranges
            
            for row_idx, row in df.iterrows():
                f.write('<tr>\n')
                for col_idx, cell in enumerate(row):
                    cell_value = '' if pd.isna(cell) else cell

                    # 检查单元格是否是合并单元格的起始单元格
                    cell_ref = sheet.cell(row=row_idx+1, column=col_idx+1).coordinate
                    merge_range = None
                    for mr in merged_cells:
                        if cell_ref in mr:
                            merge_range = mr
                            break
                    
                    if merge_range and cell_ref == merge_range.coord.split(':')[0]:
                        rowspan = merge_range.size['rows']
                        colspan = merge_range.size['columns']
                        f.write(f'<td rowspan="{rowspan}" colspan="{colspan}">{cell_value}</td>\n')
                    elif not merge_range:
                        f.write(f'<td>{cell_value}</td>\n')
                f.write('</tr>\n')
            f.write('</table>\n\n')

if __name__ == "__main__":
    outputpath = filepath + '.md'      # 输出文件
    excel_to_markdown_html(input_file = filepath, output_file = outputpath)
    print('文件已生成，路径：',outputpath)
    input("按任意键退出...")
